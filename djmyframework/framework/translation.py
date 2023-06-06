# -*- coding: utf-8 -*-
# @Time : 2020-06-05 09:26
# @Author : xzr
# @File : translation.py
# @Software: PyCharm
# @Contact : xzregg@gmail.com
# @Desc :

from django.utils.translation import lazy
from django.utils.translation import gettext as _gettext

from django.utils.translation import get_language, activate

from openpyxl import Workbook, load_workbook
import os.path

# _ = gettext_lazy

_trans_map = {}


def get_excel_text(message):
    if not _trans_map:
        load_locale_execl_file()
    return _trans_map.get(get_language(), {}).get(message, _gettext(message))


gettext = lazy(get_excel_text, str)
_ = gettext


def load_locale_execl_file():
    from django.conf import settings
    for locale_path in settings.LOCALE_PATHS:
        for languages_code, languages_name in settings.LANGUAGES:
            languages_excel_file = os.path.join(locale_path, languages_code, 'LC_MESSAGES', 'django_%s.xlsx' % languages_code)
            if os.path.exists(languages_excel_file):
                read_trans_from_excel(languages_code, languages_excel_file)


def read_trans_from_excel(languages_code, excel_file):
    global _trans_map
    _trans_map.setdefault(languages_code, {})
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        if row[0] and row[1]:
            _trans_map[languages_code][row[0]] = row[1]
    workbook.close()
