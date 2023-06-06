import os

from django.core.management.commands.makemessages import Command as MakeMessagesCommand
from django.conf import settings

import polib
from openpyxl import Workbook, load_workbook


class Command(MakeMessagesCommand):
    help = (
            """重写  makemessages  -i venv -i tests -l en"""
    )
    missing_args_message = "You must provide an application name."

    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument(
                '--merge', '-m', action='store_true',
                help='合并 xlsx 到 po 文件',
        )

    def handle(self, **options):
        self.is_merge = options['merge']
        options['ignore_patterns'].extend(['tests', 'venv', 'venv'])
        super().handle(**options)

    def get_xlsx_path(self, locale, basedir):
        return os.path.join(basedir, '%s_%s.xlsx' % (self.domain, locale))

    def merge_xlsx_file_to_po(self, pofile, locale, basedir):
        excel_file = self.get_xlsx_path(locale, basedir)
        if os.path.isfile(excel_file):
            wb = load_workbook(excel_file)
            po = polib.pofile(pofile)
            worksheet = wb.worksheets[0]
            for cell in worksheet:
                msgid = cell[0].value
                msgstr = cell[1].value
                entry = po.find(msgid)
                if entry:
                    if entry.msgid_plural:
                        entry.msgstr_plural['0'] = msgstr
                    entry.msgstr = msgstr
                    if 'fuzzy' in entry.flags:
                        entry.flags.remove('fuzzy')
            po.save()

    def write_xlsx(self, pofile, locale, basedir):
        """重写po到 excel文件"""

        wb = Workbook()
        ws = wb.active

        excel_file = self.get_xlsx_path(locale, basedir)
        po = polib.pofile(pofile)

        for entry in po:
            if not entry.obsolete:
                ws.append([entry.msgid, entry.msgstr])

        wb.save(excel_file)

    def write_po_file(self, potfile, locale):
        super().write_po_file(potfile, locale)
        basedir = os.path.join(os.path.dirname(potfile), locale, 'LC_MESSAGES')
        os.makedirs(basedir, exist_ok=True)
        pofile = os.path.join(basedir, '%s.po' % self.domain)
        if self.is_merge:
            self.merge_xlsx_file_to_po(pofile, locale, basedir)
        else:
            self.write_xlsx(pofile, locale, basedir)
